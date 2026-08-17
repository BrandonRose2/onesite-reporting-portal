from pathlib import Path
from difflib import SequenceMatcher
import json
import re

from openpyxl import load_workbook


ACTIVE_PROPERTIES = {
    "135th Street Apartments": 8, "Anaheim Gardens": 1, "Arbor Crest": 22,
    "Bayou Pointe": 30030, "Boca Ciega Townhomes": 5, "Breckenridge Village": 14,
    "Coral Village": 26, "Crossroads of Lees Summit": 30035, "Cumberland Apartments": 11,
    "Fairfax Sr Apartments": 2, "Grace Townhomes": 12, "Granite Ridge Apartments": 15,
    "Grove Park Terrace": 25, "Holiday Apartments": 10, "Howell Place": 30031,
    "Jefferson Arms Apts": 6, "Lexington Arms": 13, "Macedonia Gardens": 7,
    "Marrero 3 LP": 21, "Midtown Manor": 3, "New Wilmington Arms": 9,
    "North Pointe": 27, "Pacific Pointe Apartments": 16, "Pelican Bay": 30032,
    "Pirates Bend": 30033, "Riverchase": 17, "Silver Springs Terrace": 18,
    "St Charles": 23, "Thomasville Church Homes": 24, "Urban Rehab": 4,
    "Walnut Hill": 30034, "Windsor Village": 19, "Yorkshire Apartments": 20,
}

ALIASES = {
    "fairfaxsr apts": "Fairfax Sr Apartments",
    "fairfaxsrapartments": "Fairfax Sr Apartments",
    "jeffersonarms": "Jefferson Arms Apts",
    "pacificpointe": "Pacific Pointe Apartments",
    "silverspringsterraceapts": "Silver Springs Terrace",
    "walnuthillapartments": "Walnut Hill",
    "cumberlandapts": "Cumberland Apartments",
    "historicalriverchasehomes": "Riverchase",
    "holidayapts": "Holiday Apartments",
    "macedoniagardensapts": "Macedonia Gardens",
    "arborcrestapartments": "Arbor Crest",
    "pacificpointeapts": "Pacific Pointe Apartments",
    "yorkshireapts": "Yorkshire Apartments",
    "stcharlesapartments": "St Charles",
    "groveparkterraceapts": "Grove Park Terrace",
    "bocaciega": "Boca Ciega Townhomes",
}


def norm(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


source = Path('/home/ubuntu/upload/CompanyContacts7.23.26.xlsx')
workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook['Properties']
rows = list(sheet.iter_rows(values_only=True))
headers = [str(value).strip() if value is not None else '' for value in rows[0]]
active_by_key = {norm(name): name for name in ACTIVE_PROPERTIES}

contacts = []
for values in rows[1:]:
    record = {headers[index]: values[index] for index in range(len(headers))}
    property_name = str(record.get('Property') or '').strip()
    if not property_name or property_name.lower() in {'assistant', 'yardi properties'}:
        continue
    candidate = active_by_key.get(norm(property_name))
    candidate = ALIASES.get(norm(property_name), candidate)
    score = 1.0 if candidate else 0.0
    if not candidate:
        best_name, score = max(
            ((name, SequenceMatcher(None, norm(property_name), norm(name)).ratio()) for name in ACTIVE_PROPERTIES),
            key=lambda item: item[1],
        )
        candidate = best_name if score >= 0.82 else None
    contacts.append({
        'sourceProperty': property_name,
        'propertyName': candidate,
        'propertyId': ACTIVE_PROPERTIES.get(candidate) if candidate else None,
        'matchScore': round(score, 3),
        'managerName': record.get('Manager'),
        'officePhone': str(record.get('Office') or '').strip() or None,
        'extension': str(record.get('Ext') or '').strip() or None,
        'mobilePhone': str(record.get('Mobile') or '').strip() or None,
        'managerEmail': str(record.get('Email Address') or '').strip() or None,
    })

matched = [item for item in contacts if item['propertyId']]
review_required = [item for item in contacts if not item['propertyId']]
result = {
    'matched': matched,
    'reviewRequired': review_required,
    'summary': {
        'matched': len(matched),
        'withExtensions': sum(1 for item in matched if item['extension']),
        'reviewRequired': len(review_required),
    },
}

def sql_literal(value):
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''") + "'"

updates = []
for item in matched:
    updates.append(
        "UPDATE propertyContacts SET "
        f"managerName={sql_literal(item['managerName'])}, "
        f"managerEmail={sql_literal(item['managerEmail'])}, "
        f"mobilePhone={sql_literal(item['mobilePhone'])}, "
        f"officePhone={sql_literal(item['officePhone'])}, "
        f"extension={sql_literal(item['extension'])}, "
        f"sourcePropertyName={sql_literal(item['sourceProperty'])}, "
        "sourcePageTitle='Company Contacts 7.23.26', "
        "sourceUrl=NULL, mappingStatus='verified', sourceSyncedAt=CURRENT_TIMESTAMP "
        f"WHERE propertyId={item['propertyId']};"
    )

Path('/home/ubuntu/company_contacts_import.sql').write_text("\n".join(updates) + "\n", encoding='utf-8')
print(json.dumps(result, indent=2, default=str))
